import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import io
import re
import json
from google import genai
from google.genai import types

st.set_page_config(page_title="AI Data Agent Engine", layout="wide")
st.title("🤖 Autonomous AI Data Agent & Visualizer")
st.write("Upload any chaotic business sheet. The AI dynamically understands context, cleans anomalies, maps visualizations, and drafts analytical briefs.")

# Initialize the Gemini Client securely
# Tip: Go to settings in Streamlit / Secrets and add: GEMINI_API_KEY = "your_key"
# If not set, it gracefully defaults to user input or falls back to rule-based classification
api_key = st.sidebar.text_input("Enter Gemini API Key (Optional)", type="password") or st.secrets.get("GEMINI_API_KEY")

if not api_key:
    st.sidebar.warning("⚠️ Enter a Gemini API Key to unlock AI Insights and smart schema understanding! Falling back to Rule Engine mode.")

# File Upload Widget
uploaded_file = st.file_uploader("Choose a messy Excel or CSV file", type=['xlsx', 'csv'])

if uploaded_file is not None:
    try:
        # Load and clean up raw file structures
        if uploaded_file.name.endswith('.csv'):
            raw_bytes = uploaded_file.getvalue()
            raw_text = raw_bytes.decode('utf-8', errors='ignore')
            lines = raw_text.splitlines()
            
            header_index = 0
            for i, line in enumerate(lines):
                cleaned_line = line.strip().upper()
                matches = sum(1 for k in ['ID', 'INVOICE', 'NAME', 'DATE', 'PRICE', 'AMOUNT', 'CATEGORY', 'COUNTRY'] if k in cleaned_line)
                if matches >= 2:
                    header_index = i
                    break
            
            valid_lines = lines[header_index:]
            clean_csv_data = "\n".join(valid_lines)
            df = pd.read_csv(io.StringIO(clean_csv_data), on_bad_lines='skip')
        else:
            df = pd.read_excel(uploaded_file)

        # -------------------------------------------------------------
        # CORE CLEANING PIPELINE
        # -------------------------------------------------------------
        df = df.dropna(how='all')
        df.columns = [str(col).strip() for col in df.columns]
        df = df.loc[:, ~df.columns.str.contains('^Unnamed', case=False, na=False)]
        df.columns = [col.title() for col in df.columns]
        
        # Robust multi-format date parser engine
        def robust_date_parser(val):
            if pd.isna(val):
                return None
            val_clean = str(val).strip().replace('.', '-').replace('/', '-')
            if val_clean.lower() in ['none', 'nan', 'nat', '', 'n/a']:
                return None
            for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%m-%d-%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%y', '%m-%d-%y']:
                try:
                    return pd.to_datetime(val_clean, format=fmt).strftime('%Y-%m-%d')
                except:
                    continue
            try:
                return pd.to_datetime(val_clean, errors='coerce', dayfirst=True).strftime('%Y-%m-%d')
            except:
                return None

        # Base rule classifications (Fallback)
        numeric_cols = [col for col in df.columns if any(k in col.lower() for k in ['price', 'amount', 'sales', 'revenue', 'cost'])]
        category_cols = [col for col in df.columns if any(k in col.lower() for k in ['category', 'item', 'product', 'country', 'location', 'status'])]
        date_cols = [col for col in df.columns if 'date' in col.lower()]

        # --- ADVANCED AI SCHEMA INTELLIGENCE INTERFACE ---
        if api_key:
            try:
                client = genai.Client(api_key=api_key)
                sample_data = df.head(5).to_json()
                
                ai_prompt = f"""
                Analyze the columns of this business dataset and classify them into exactly three logical segments.
                Columns list: {list(df.columns)}
                Sample Data: {sample_data}
                
                Respond ONLY with a valid JSON object matching this structure:
                {{
                    "numeric_columns": ["list", "of", "financial_or_numeric_metric_columns"],
                    "category_columns": ["list", "of", "categorical_grouping_columns"],
                    "date_columns": ["list", "of", "timestamp_or_date_columns"]
                }}
                """
                
                response = client.models.generate_content(
                    model='gemini-2.5-flash',
                    contents=ai_prompt,
                    config=types.GenerateContentConfig(response_mime_type="application/json")
                )
                
                ai_schema = json.loads(response.text)
                # Verify AI results against existing frame arrays
                numeric_cols = [c for c in ai_schema.get('numeric_columns', []) if c in df.columns]
                category_cols = [c for c in ai_schema.get('category_columns', []) if c in df.columns]
                date_cols = [c for c in ai_schema.get('date_columns', []) if c in df.columns]
            except Exception as e:
                st.sidebar.error(f"AI Classification Error: {e}. Reverting to rule matrices.")

        # Standardize matching columns based on structured target schemas
        for col in df.columns:
            if col in date_cols:
                df[col] = df[col].apply(robust_date_parser)
            elif col in numeric_cols:
                df[col] = df[col].astype(str).str.replace(r'[₹$,\s]', '', regex=True)
                df[col] = pd.to_numeric(df[col], errors='coerce')
            else:
                if df[col].dtype == 'object':
                    df[col] = df[col].astype(str).str.strip()
                    df[col] = df[col].replace({'nan': None, 'N/A': None, 'None': None, 'NaN': None, '': None})

        df = df.dropna(how='all')

        # -------------------------------------------------------------
        # EXECUTIVE MANAGEMENT USER INTERFACE
        # -------------------------------------------------------------
        st.write("---")
        st.subheader("📈 Executive Performance Summary")
        
        kpi1, kpi2, kpi3, kpi4 = st.columns(4)
        kpi1.metric("Total Records", f"{len(df)}")
        if numeric_cols:
            kpi2.metric(f"Total {numeric_cols[0]}", f"₹{df[numeric_cols[0]].sum():,.2f}")
            kpi3.metric(f"Average {numeric_cols[0]}", f"₹{df[numeric_cols[0]].mean():,.2f}")
        else:
            kpi2.metric("Financial Metrics", "No Data")
            kpi3.metric("Average Transaction", "N/A")
        if category_cols:
            kpi4.metric(f"Unique {category_cols[0]}s", f"{df[category_cols[0]].nunique()}")

        # Generate automated narrative insights brief using AI
        if api_key and numeric_cols:
            with st.spinner("🤖 Autonomous AI Data Agent is drafting your summary brief..."):
                try:
                    summary_stats = df.describe(include='all').to_string()
                    brief_prompt = f"Act as an Elite Business Performance Analyst. Draft a concise, high-impact 3-sentence executive analytical brief based on this data summary layout: {summary_stats}. Focus on notable performance drivers."
                    
                    brief_response = client.models.generate_content(
                        model='gemini-2.5-flash',
                        contents=brief_prompt
                    )
                    st.info(f"**AI Analyst Insights:** {brief_response.text}")
                except Exception as e:
                    pass

        tab1, tab2 = st.tabs(["👀 Cleaned Data View", "📊 Graphical Visualizations"])
        with tab1:
            st.dataframe(df, use_container_width=True)
        with tab2:
            chart_col1, chart_col2 = st.columns(2)
            with chart_col1:
                if category_cols and numeric_cols:
                    chart_data = df.groupby(category_cols[0])[numeric_cols[0]].sum().reset_index()
                    st.bar_chart(data=chart_data, x=category_cols[0], y=numeric_cols[0], use_container_width=True)
            with chart_col2:
                if date_cols and numeric_cols:
                    trend_data = df.groupby(date_cols[0])[numeric_cols[0]].sum().reset_index().sort_values(by=date_cols[0])
                    st.line_chart(data=trend_data, x=date_cols[0], y=numeric_cols[0], use_container_width=True)

        # -------------------------------------------------------------
        # ADVANCED EXCEL STYLING & EMBEDDED CHART ENGINE
        # -------------------------------------------------------------
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Clean Data', index=False)
            workbook = writer.book
            data_sheet = workbook['Clean Data']
            
            dashboard_sheet = workbook.create_sheet(title="Executive Dashboard", index=0)
            dashboard_sheet.views.sheetView[0].showGridLines = True
            
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') 
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            data_font = Font(name='Calibri', size=11)
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')
            thin_side = Side(border_style="thin", color="D9D9D9")
            cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            for col_num in range(1, len(df.columns) + 1):
                cell = data_sheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            
            for row in range(2, data_sheet.max_row + 1):
                for col in range(1, data_sheet.max_column + 1):
                    cell = data_sheet.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = cell_border
                    val_str = str(cell.value or '').lower()
                    
                    if isinstance(cell.value, (int, float)) and not (data_sheet.cell(row=1, column=col).value in ['Phone Number', 'Phone', 'Mobile']):
                        cell.alignment = right_align
                        cell.number_format = '#,##0.00'
                    elif re.match(r'^\d{4}-\d{2}-\d{2}$', val_str) or 'inv-' in val_str or 'trx-' in val_str:
                        cell.alignment = center_align
                    else:
                        if data_sheet.cell(row=1, column=col).value in ['Phone Number', 'Phone', 'Mobile']:
                            cell.number_format = '@'
                        cell.alignment = left_align

            for col in data_sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                data_sheet.column_dimensions[col_letter].width = max(max_len + 4, 13)

            # --- GENERATING NATIVE EXCEL CHARTS SAFELY ---
            if category_cols and numeric_cols:
                dashboard_sheet["A2"] = "KPI Dashboard Summary"
                dashboard_sheet["A2"].font = Font(name='Calibri', size=16, bold=True, color='1F4E78')
                
                dashboard_sheet["A4"] = category_cols[0]
                dashboard_sheet["B4"] = numeric_cols[0]
                dashboard_sheet["A4"].font = header_font
                dashboard_sheet["A4"].fill = header_fill
                dashboard_sheet["B4"].font = header_font
                dashboard_sheet["B4"].fill = header_fill
                
                summary_data = df.groupby(category_cols[0])[numeric_cols[0]].sum().reset_index()
                for i, row_data in summary_data.iterrows():
                    r = 5 + i
                    dashboard_sheet.cell(row=r, column=1, value=row_data[category_cols[0]]).alignment = left_align
                    val_cell = dashboard_sheet.cell(row=r, column=2, value=row_data[numeric_cols[0]])
                    val_cell.number_format = '#,##0.00'
                    val_cell.alignment = right_align
                
                chart1 = BarChart()
                chart1.type = "col"
                chart1.style = 10
                chart1.title = f"Total {numeric_cols[0]} by {category_cols[0]}"
                chart1.y_axis.title = numeric_cols[0]
                chart1.x_axis.title = category_cols[0]
                
                data_ref = Reference(dashboard_sheet, min_col=2, min_row=4, max_row=4+len(summary_data))
                cats_ref = Reference(dashboard_sheet, min_col=1, min_row=5, max_row=4+len(summary_data))
                
                chart1.add_data(data_ref, titles_from_data=True)
                chart1.set_categories(cats_ref)
                chart1.height = 14
                chart1.width = 22
                dashboard_sheet.add_chart(chart1, "D4")
            else:
                # Safe fallback if file structure has text categories only
                if category_cols:
                    dashboard_sheet["A2"] = f"Distribution Frequency Summary of {category_cols[0]}"
                    dashboard_sheet["A2"].font = Font(name='Calibri', size=14, bold=True, color='1F4E78')
                    
                    dashboard_sheet["A4"] = category_cols[0]
                    dashboard_sheet["B4"] = "Incident/Record Frequency Count"
                    dashboard_sheet["A4"].font = header_font
                    dashboard_sheet["A4"].fill = header_fill
                    dashboard_sheet["B4"].font = header_font
                    dashboard_sheet["B4"].fill = header_fill
                    
                    summary_data = df[category_cols[0]].value_counts().reset_index()
                    for i, row_data in summary_data.iterrows():
                        r = 5 + i
                        dashboard_sheet.cell(row=r, column=1, value=row_data.iloc[0]).alignment = left_align
                        val_cell = dashboard_sheet.cell(row=r, column=2, value=int(row_data.iloc[1]))
                        val_cell.alignment = right_align
                    
                    chart1 = BarChart()
                    chart1.type = "col"
                    chart1.title = f"Volume Breakdown of {category_cols[0]}"
                    data_ref = Reference(dashboard_sheet, min_col=2, min_row=4, max_row=4+len(summary_data))
                    cats_ref = Reference(dashboard_sheet, min_col=1, min_row=5, max_row=4+len(summary_data))
                    chart1.add_data(data_ref, titles_from_data=True)
                    chart1.set_categories(cats_ref)
                    chart1.height = 14
                    chart1.width = 22
                    dashboard_sheet.add_chart(chart1, "D4")
                else:
                    dashboard_sheet["A2"] = "Upload a standardized sheet with clean data descriptors to automatically render layout charts."

        processed_data = output.getvalue()
        
        st.write("---")
        st.success("🎉 Processed data metrics and native charts ready for download!")
        st.download_button(
            label="📥 Download Beautiful Clean Sheet with Charts",
            data=processed_data,
            file_name=f"Cleaned_{uploaded_file.name if uploaded_file.name.endswith('.xlsx') else 'Data.xlsx'}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        st.error(f"Error processing file: {e}")
